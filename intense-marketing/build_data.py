# -*- coding: utf-8 -*-
"""
Solveur + exporteur pour le jeu de déduction recrutement.
Lit jeu_deduction_recrutement.xlsx, calcule la réponse de chaque fiche,
vérifie l'unicité de la solution, et exporte data.js pour le jeu web.
"""
import sys, io, json, itertools
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

XLSX = 'jeu_deduction_recrutement.xlsx'

# ─── Candidats (id 1..7) ────────────────────────────────────────────────
# exp = expérience en années (= "taille"), mode = mode de travail
CANDIDATES = {
    1: {"animal": "Panthère", "role": "Ingénieur",          "exp": 7,  "mode": "Remote", "name": "Thomas Leroy",    "img": "characters/1 - Ingenieur.png"},
    2: {"animal": "Paon",     "role": "Designer",           "exp": 3,  "mode": "Bureau", "name": "Camille Dupont",  "img": "characters/2 - Designer.png"},
    3: {"animal": "Requin",   "role": "Commercial",         "exp": 11, "mode": "Remote", "name": "Julien Martin",   "img": "characters/3 - Commercial.png"},
    4: {"animal": "Iguane",   "role": "Développeur",        "exp": 1,  "mode": "Bureau", "name": "Maxime Giraud",   "img": "characters/4 - Developpeur.png"},
    5: {"animal": "Chien",    "role": "Manager",            "exp": 5,  "mode": "Remote", "name": "Sophie Leroy",    "img": "characters/5 - Manager.png"},
    6: {"animal": "Cheval",   "role": "Recrutement Humain", "exp": 15, "mode": "Bureau", "name": "Isabelle Moreau", "img": "characters/6 - Recrutement Humain.png"},
    7: {"animal": "Ours",     "role": "Data Analyst",       "exp": 9,  "mode": "Remote", "name": "Marcus Diop",     "img": "characters/7 - Data Analyst.png"},
}

# ─── Définition logique de chaque carte ─────────────────────────────────
# Une fonction card(id) -> predicate(active, comp:set, liar:set, holder) -> bool
# active : liste ordonnée des ids actifs (= ordre des joueurs)
# comp   : set des ids concurrents (coupables)
# liar   : set des ids menteurs
# holder : id du candidat qui porte la carte (pour les cartes "moi")

E = lambda i: CANDIDATES[i]["exp"]
M = lambda i: CANDIDATES[i]["mode"]

def neighbors(active, h):
    """voisins directs de h — disposition CIRCULAIRE (le 1er voisin du dernier)"""
    if h not in active: return []
    n = len(active)
    if n < 2: return []
    i = active.index(h)
    return list({active[(i-1) % n], active[(i+1) % n]})

def between(active, x):
    """(gauche, droite) en cercle ; None si moins de 3 joueurs"""
    if x not in active: return None
    n = len(active)
    if n < 3: return None
    i = active.index(x)
    return (active[(i-1) % n], active[(i+1) % n])

def adj_pairs(active):
    """paires de voisins consécutifs en cercle"""
    n = len(active)
    if n < 2: return []
    if n == 2: return [(active[0], active[1])]
    return [(active[i], active[(i+1) % n]) for i in range(n)]

def adj_triples(active):
    """triplets de voisins consécutifs en cercle"""
    n = len(active)
    if n < 3: return []
    return [(active[i], active[(i+1) % n], active[(i+2) % n]) for i in range(n)]

CARDS = {
  1:  lambda a,c,l,h: h in c,
  2:  lambda a,c,l,h: h not in c,
  3:  lambda a,c,l,h: 1 in c,
  4:  lambda a,c,l,h: 3 in c,
  5:  lambda a,c,l,h: 5 in c,
  6:  lambda a,c,l,h: 7 in c,
  7:  lambda a,c,l,h: 2 not in c,
  8:  lambda a,c,l,h: 4 not in c,
  9:  lambda a,c,l,h: 6 not in c,
  10: lambda a,c,l,h: 2 in l,
  11: lambda a,c,l,h: 3 in l,
  12: lambda a,c,l,h: 5 in l,
  13: lambda a,c,l,h: 6 in l,
  14: lambda a,c,l,h: 1 in l and 2 in l,
  15: lambda a,c,l,h: 3 in l and 4 in l,
  16: lambda a,c,l,h: 6 in l and 7 in l,
  17: lambda a,c,l,h: 2 not in c and 3 not in c,
  18: lambda a,c,l,h: 4 not in c and 5 not in c,
  19: lambda a,c,l,h: 1 in c or 2 in c,
  20: lambda a,c,l,h: 2 in c or 3 in c,
  21: lambda a,c,l,h: 6 in c or 7 in c,
  22: lambda a,c,l,h: 3 not in l and 4 not in l,
  23: lambda a,c,l,h: 5 not in l and 6 not in l,
  24: lambda a,c,l,h: 2 not in c and h not in c,
  25: lambda a,c,l,h: 5 not in c and h not in c,
  26: lambda a,c,l,h: 1 in c or h in c,
  27: lambda a,c,l,h: 6 in c or h in c,
  28: lambda a,c,l,h: any(E(x) > E(h) and x in c for x in a),
  29: lambda a,c,l,h: any(E(x) < E(h) and x in c for x in a),
  30: lambda a,c,l,h: any(E(x) > 6  and x in c for x in a),
  31: lambda a,c,l,h: any(E(x) > 10 and x in c for x in a),
  32: lambda a,c,l,h: any(E(x) < 4  and x in c for x in a),
  33: lambda a,c,l,h: any(E(x) < 8  and x in c for x in a),
  34: lambda a,c,l,h: any(x in c and M(x) == "Remote" for x in a),
  35: lambda a,c,l,h: any(x in c and M(x) == "Bureau" for x in a),
  36: lambda a,c,l,h: any(n in c for n in neighbors(a, h)),
  37: lambda a,c,l,h: all(n not in c for n in neighbors(a, h)),
  38: lambda a,c,l,h: any(x in c and x in l for x in a),
  39: lambda a,c,l,h: any(x in c and x not in l for x in a),
  40: lambda a,c,l,h: any((between(a,x) and x in c and between(a,x)[0] in l and between(a,x)[1] in l) for x in a),
  41: lambda a,c,l,h: any(x in l and y in l for x,y in adj_pairs(a)),
  42: lambda a,c,l,h: (max((x for x in a if x in l), key=E, default=None) in c) if any(x in l for x in a) else False,
  43: lambda a,c,l,h: sum(1 for x in a if x in c and x in l) >= 2,
  44: lambda a,c,l,h: any((between(a,x) and x in c and between(a,x)[0] not in c and between(a,x)[1] not in c) for x in a),
  45: lambda a,c,l,h: any(x in c and y in c for x,y in adj_pairs(a)),
  46: lambda a,c,l,h: len({M(x) for x in a if x in c}) >= 2,
  47: lambda a,c,l,h: any(x in c and x in l for x in a) and any(x in c and x not in l for x in a),
  48: lambda a,c,l,h: any(x not in c and y not in c and z not in c for x,y,z in adj_triples(a)),
  49: lambda a,c,l,h: (min((x for x in a if x in c), key=E, default=None) not in l) if any(x in c for x in a) else False,
  50: lambda a,c,l,h: any(abs(E(x)-E(y)) <= 4 for x,y in itertools.combinations([z for z in a if z in c], 2)),
}

# ─── Solveur ────────────────────────────────────────────────────────────
def solve(active, hand, n_comp, n_liar):
    """
    active : liste ordonnée des ids actifs
    hand   : dict id -> card_id (carte portée par chaque candidat)
    Retourne la liste des états valides (frozenset comp, frozenset liar).
    """
    solutions = []
    ids = list(active)
    for comp in itertools.combinations(ids, n_comp):
        comp = set(comp)
        for liar in itertools.combinations(ids, n_liar):
            liar = set(liar)
            ok = True
            for h in ids:
                card = CARDS[hand[h]]
                truth = card(active, comp, liar, h)
                # menteur => carte fausse ; non-menteur => carte vraie
                expected_true = (h not in liar)
                if truth != expected_true:
                    ok = False
                    break
            if ok:
                solutions.append((frozenset(comp), frozenset(liar)))
    return solutions

# ─── Lecture des fiches ─────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Fiches mini-jeux']
rows = list(ws.iter_rows(values_only=True))

fiches = []
for row in rows[1:]:
    if row[0] is None: continue
    n_fiche, n_jeu, n_players, n_comp, n_liar = (int(row[i]) for i in range(5))
    cards = [int(row[5+i]) for i in range(n_players)]
    active = list(range(1, n_players+1))       # candidats 1..N (convention vérifiée)
    hand = {active[i]: cards[i] for i in range(n_players)}
    sols = solve(active, hand, n_comp, n_liar)

    comp_sets = {s[0] for s in sols}
    unique_comp = (len(comp_sets) == 1)
    unique_full = (len(sols) == 1)
    answer = sorted(next(iter(comp_sets))) if unique_comp else None

    fiches.append({
        "fiche": n_fiche, "jeu": n_jeu,
        "players": n_players, "n_comp": n_comp, "n_liar": n_liar,
        "active": active, "hand": cards,
        "n_solutions": len(sols),
        "unique_comp": unique_comp, "unique_full": unique_full,
        "answer": answer,
        "answer_full": [sorted(answer)] if (answer and unique_full) else None,
        "liars": sorted(next(iter(sols))[1]) if unique_full else None,
    })

# ─── Rapport ────────────────────────────────────────────────────────────
total = len(fiches)
uniq_c = sum(1 for f in fiches if f["unique_comp"])
uniq_f = sum(1 for f in fiches if f["unique_full"])
no_sol = sum(1 for f in fiches if f["n_solutions"] == 0)
print(f"Total fiches            : {total}")
print(f"Solution concurrents unique : {uniq_c}/{total}")
print(f"Solution complète unique    : {uniq_f}/{total}")
print(f"Aucune solution             : {no_sol}/{total}")
print()
print("Détail des fiches NON-uniques (concurrents) :")
for f in fiches:
    if not f["unique_comp"]:
        print(f"  fiche {f['fiche']} jeu {f['jeu']}: {f['n_solutions']} sol, "
              f"cartes={f['hand']}, comp={f['n_comp']}, liar={f['n_liar']}")

# Aperçu de quelques réponses
print("\nAperçu (10 premières) :")
for f in fiches[:10]:
    roles = [CANDIDATES[i]['role'] for i in (f['answer'] or [])]
    print(f"  fiche {f['fiche']} jeu {f['jeu']}: concurrents = {f['answer']} {roles}")

# ─── Lecture des textes de cartes ───────────────────────────────────────
ws_c = wb['Cartes']
card_texts = {}
for row in list(ws_c.iter_rows(values_only=True))[1:]:
    if row[0] is None: continue
    cid = int(row[0])
    card_texts[cid] = {
        "original":   (row[1] or "").strip(),
        "translated": (row[2] or "").strip(),
    }

# ─── Export data.js ─────────────────────────────────────────────────────
# On ne garde que les fiches à solution (concurrents) unique = toutes ici.
export = {
    "candidates": CANDIDATES,
    "cards": card_texts,
    "fiches": [
        {
            "id": f"{f['fiche']}-{f['jeu']}",
            "fiche": f["fiche"], "jeu": f["jeu"],
            "players": f["players"],
            "n_comp": f["n_comp"], "n_liar": f["n_liar"],
            "active": f["active"],
            "hand": f["hand"],                 # carte de chaque joueur (ordre = cercle)
            "answer": f["answer"],             # ids des concurrents (solution)
        }
        for f in fiches if f["unique_comp"]
    ],
}

with io.open("data.js", "w", encoding="utf-8") as fp:
    fp.write("// Généré par build_data.py — ne pas éditer à la main.\n")
    fp.write("window.GAME_DATA = ")
    json.dump(export, fp, ensure_ascii=False, indent=2)
    fp.write(";\n")

print(f"\n✅ data.js exporté : {len(export['fiches'])} fiches, "
      f"{len(card_texts)} cartes, {len(CANDIDATES)} candidats.")
